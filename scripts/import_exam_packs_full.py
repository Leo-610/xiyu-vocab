#!/usr/bin/env python3
"""
完整导入专四/专八：从微信压缩包抽取 Excel+图片，与现有义项包合并入库。

压缩包内 Excel：
  images专业四级2.zip → senses_table_tem4 7.16.xlsx (+143 图)
  images专业四级3.zip → senses_table_tem4 7.17.xlsx (+143 图)
  images专业四级4.zip → senses_table_tem4 7.18.xlsx (+163 图)
  images专业四级5.zip → senses_table_tem4 7.19.xlsx (+170 图)
  images专业四级6.zip → senses_table_tem4 7.20.xlsx
  images专业八级1.zip → 59 图（词表用 senses_table_tem8_with_images.xlsx）

用法:
  python3 scripts/import_exam_packs_full.py
  node scripts/reset-with-exams.mjs
"""

from __future__ import annotations

import csv
import random
import re
import shutil
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from PIL import Image
except ImportError:
    raise SystemExit("请先安装: pip3 install pillow openpyxl")

ROOT = Path(__file__).resolve().parent.parent
CONTENT = ROOT / "data" / "content"
IMAGES = ROOT / "data" / "images"
BATCHES = ROOT / "data" / "batches"
STAGING = ROOT / "data" / "content" / "_exam_staging"

WECHAT = Path(
    "/Users/liuyiming/Library/Containers/com.tencent.xinWeChat/"
    "Data/Documents/xwechat_files/wxid_56yn1iomktsr12_9b4c/msg/file/2026-07"
)

ZIPS = [
    ("tem4_716", WECHAT / "images专业四级2.zip", "专四"),
    ("tem4_717", WECHAT / "images专业四级3.zip", "专四"),
    ("tem4_718", WECHAT / "images专业四级4.zip", "专四"),
    ("tem4_719", WECHAT / "images专业四级5.zip", "专四"),
    ("tem4_720", WECHAT / "images专业四级6.zip", "专四"),
    ("tem8", WECHAT / "images专业八级1(1).zip", "专八"),
]

TEM4_BASE_XLSX = [
    WECHAT / "senses_table_tem4 7.15(1).xlsx",
    CONTENT / "senses_table_tem4.xlsx",
]
TEM4_WITH_IMG = WECHAT / "senses_table_tem4_with_images.xlsx"
TEM8_WITH_IMG = WECHAT / "senses_table_tem8_with_images.xlsx"
A1_CSV = BATCHES / "A1" / "words_senses_team.csv"

CSV_FIELDS = [
    "lemma", "pos", "gender", "level", "sense", "ipa",
    "meaning_zh", "meaning_en", "example_es", "example_zh",
    "image_file", "tags", "exam_tags",
    "distractor_1", "distractor_2", "distractor_3", "copyright_note",
]

CAT_MAP = {
    "f": ("n", "f"),
    "m": ("n", "m"),
    "m, f": ("n", "n/a"),
    "f, m": ("n", "n/a"),
    "v": ("v", "n/a"),
    "tr": ("v", "n/a"),
    "intr": ("v", "n/a"),
    "tr, intr": ("v", "n/a"),
    "intr, tr": ("v", "n/a"),
    "prnl": ("v", "n/a"),
    "adj": ("adj", "n/a"),
    "adv": ("adv", "n/a"),
    "interj": ("int", "n/a"),
    "int": ("int", "n/a"),
    "pron": ("pron", "n/a"),
    "prep": ("prep", "n/a"),
    "conj": ("conj", "n/a"),
}


def safe_copy(src: Path, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists():
        dest.chmod(0o644)
        dest.unlink()
    shutil.copy2(src, dest)
    dest.chmod(0o644)


def slugify(text: str) -> str:
    s = unicodedata.normalize("NFD", str(text).strip().lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_]+", "_", s)
    return s[:80] or "word"


def map_categoria(cat: str) -> tuple[str, str]:
    key = re.sub(r"\s+", " ", (cat or "").strip().lower())
    return CAT_MAP.get(key, ("n", "n/a"))


def normalize_media_target(target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("../media/"):
        return "xl/media/" + target.split("media/", 1)[1]
    if target.startswith("/xl/media/"):
        return target.lstrip("/")
    if target.startswith("media/"):
        return "xl/" + target
    return target


def extract_image_map(xlsx: Path) -> dict[int, str]:
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    }
    a_ns = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    r_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
    with zipfile.ZipFile(xlsx) as z:
        rel_name = "xl/drawings/_rels/drawing1.xml.rels"
        drawing_name = "xl/drawings/drawing1.xml"
        if rel_name not in z.namelist() or drawing_name not in z.namelist():
            return {}
        rel_root = ET.fromstring(z.read(rel_name))
        rid_map = {
            rel.attrib["Id"]: normalize_media_target(rel.attrib["Target"])
            for rel in rel_root
            if rel.attrib.get("Type", "").endswith("/image")
        }
        root = ET.fromstring(z.read(drawing_name))
        mapping: dict[int, str] = {}
        for anc in root:
            frm = anc.find("xdr:from", ns)
            if frm is None:
                continue
            excel_row = int(frm.findtext("xdr:row", namespaces=ns)) + 1
            blip = anc.find(f".//{a_ns}blip")
            if blip is None:
                continue
            embed = blip.attrib.get(r_attr)
            target = rid_map.get(embed)
            if target:
                mapping[excel_row] = target
    return mapping


def save_jpg(dest: Path, data: bytes) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    im = Image.open(BytesIO(data)).convert("RGB")
    if max(im.size) > 1000:
        im.thumbnail((800, 800), Image.Resampling.LANCZOS)
    im.save(dest, "JPEG", quality=82, optimize=True)


def extract_zips() -> dict[str, Path]:
    """返回 label → staging 目录（含 images/ 与 *.xlsx）。"""
    if STAGING.exists():
        shutil.rmtree(STAGING)
    STAGING.mkdir(parents=True)
    dirs: dict[str, Path] = {}
    for label, zp, _tag in ZIPS:
        if not zp.exists():
            print(f"[warn] 缺少 zip: {zp}")
            continue
        d = STAGING / label
        img_dir = d / "images"
        img_dir.mkdir(parents=True)
        with zipfile.ZipFile(zp) as zf:
            for name in zf.namelist():
                if name.endswith("/") or "__MACOSX" in name:
                    continue
                low = name.lower()
                bn = Path(name).name
                if low.endswith((".xlsx", ".xls")):
                    (d / bn).write_bytes(zf.read(name))
                    print(f"[ok] 抽出 Excel {label}/{bn}")
                elif low.endswith((".png", ".jpg", ".jpeg", ".webp")):
                    (img_dir / bn).write_bytes(zf.read(name))
        print(f"[ok] {label} 图片 {len(list(img_dir.glob('*')))} 张")
        dirs[label] = d
    return dirs


def load_xlsx_rows(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    entries = []
    for raw in rows_raw[1:]:
        if not raw or not raw[0]:
            continue
        lemma = str(raw[0]).strip()
        sense = int(raw[1] or 1)
        cat = str(raw[2] or "").strip()
        level = str(raw[3] or "A2").strip().upper() or "A2"
        meaning_zh = str(raw[4] or "").strip()
        if not meaning_zh:
            continue
        example_es = str(raw[5] or "").strip()
        example_zh = str(raw[6] or "").strip()
        image_hint = str(raw[7] or "").strip() if len(raw) > 7 and raw[7] else ""
        pos, gender = map_categoria(cat)
        if image_hint:
            stem = slugify(Path(image_hint).stem)
        else:
            stem = f"{slugify(lemma)}_{sense}"
        entries.append({
            "lemma": lemma,
            "pos": pos,
            "gender": gender,
            "level": level,
            "sense": sense,
            "meaning_zh": meaning_zh,
            "example_es": example_es,
            "example_zh": example_zh,
            "image_stem": stem,
            "cat_raw": cat,
            "source": xlsx.name,
        })
    return entries


def load_embedded_images(xlsx: Path) -> dict[tuple[str, int], bytes]:
    if not xlsx.exists():
        return {}
    image_map = extract_image_map(xlsx)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    out: dict[tuple[str, int], bytes] = {}
    with zipfile.ZipFile(xlsx) as z:
        for excel_row_idx, raw in enumerate(rows_raw[1:], start=2):
            if not raw or not raw[0]:
                continue
            lemma = str(raw[0]).strip()
            sense = int(raw[1] or 1)
            media = image_map.get(excel_row_idx)
            if media and media in z.namelist():
                out[(lemma, sense)] = z.read(media)
    print(f"[ok] 内嵌图 {xlsx.name}: {len(out)}")
    return out


def load_a1_csv() -> dict[tuple[str, int], dict]:
    rows: dict[tuple[str, int], dict] = {}
    if not A1_CSV.exists():
        return rows
    with A1_CSV.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            lemma = row["lemma"].strip()
            sense = int(row.get("sense") or 1)
            stem = slugify(Path(row["image_file"]).stem)
            rows[(lemma, sense)] = {
                "lemma": lemma,
                "pos": row.get("pos") or "n",
                "gender": row.get("gender") or "n/a",
                "level": (row.get("level") or "A1").upper(),
                "sense": sense,
                "meaning_zh": row.get("meaning_zh") or "",
                "example_es": row.get("example_es") or "",
                "example_zh": row.get("example_zh") or "",
                "image_stem": stem,
                "image_file": row.get("image_file") or f"{stem}.jpg",
                "exam_tags": set(),
                "pack": "A1",
                "tags_extra": ["义项包"],
                "source": "words_senses_team.csv",
            }
    return rows


def build_image_index(staging_dirs: dict[str, Path]) -> dict[str, Path]:
    """stem → 本地图片路径（后写覆盖）。兼容 zip 非 ASCII 文件名乱码。"""
    index: dict[str, Path] = {}

    def add(stem: str, path: Path) -> None:
        if stem:
            index[stem] = path

    for label, d in staging_dirs.items():
        img_dir = d / "images"
        if not img_dir.exists():
            continue
        for p in img_dir.iterdir():
            if p.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp"}:
                continue
            add(slugify(p.stem), p)
            # 乱码文件名常见：evaluaci¿«n → 仍尽量用已知替换
            raw = p.name
            for a, b in (
                ("¿«", "ó"),
                ("¿®", "í"),
                ("¿¡", "á"),
                ("¿©", "é"),
                ("¿║", "ú"),
                ("¿±", "ñ"),
                ("¿", ""),
            ):
                raw = raw.replace(a, b)
            add(slugify(Path(raw).stem), p)
    return index


def merge_entry(store: dict[tuple[str, int], dict], entry: dict, exam_tag: str | None) -> None:
    key = (entry["lemma"], entry["sense"])
    if key not in store:
        store[key] = {
            **entry,
            "exam_tags": set([exam_tag] if exam_tag else []),
            "pack": "exam" if exam_tag else "A1",
            "tags_extra": [],
        }
        return
    cur = store[key]
    # 新表覆盖释义/例句/等级（同学最新校对）
    for field in ("meaning_zh", "example_es", "example_zh", "pos", "gender", "level", "image_stem", "source"):
        if entry.get(field):
            cur[field] = entry[field]
    if exam_tag:
        cur["exam_tags"].add(exam_tag)
        if cur.get("pack") == "A1":
            cur["tags_extra"] = list(set(cur.get("tags_extra", []) + ["义项包"]))


def fill_distractors(entries: list[dict]) -> None:
    by_level: dict[str, list[str]] = {}
    for e in entries:
        by_level.setdefault(e["level"], []).append(e["meaning_zh"])
    rng = random.Random(42)
    for e in entries:
        pool = [m for m in by_level.get(e["level"], []) if m != e["meaning_zh"]]
        rng.shuffle(pool)
        picks = (pool + ["待填写", "待填写", "待填写"])[:3]
        e["distractor_1"], e["distractor_2"], e["distractor_3"] = picks


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_merged_xlsx(path: Path, rows: list[dict], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    headers = ["西语", "义项", "categoría", "级别", "中文释义", "例句*西语", "例句*中文", "图片", "exam_tags"]
    fill = PatternFill("solid", fgColor="1F4E79")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    for i, e in enumerate(rows, 2):
        ws.cell(i, 1, e["lemma"])
        ws.cell(i, 2, e["sense"])
        ws.cell(i, 3, e.get("cat_raw") or e["pos"])
        ws.cell(i, 4, e["level"])
        ws.cell(i, 5, e["meaning_zh"])
        ws.cell(i, 6, e["example_es"])
        ws.cell(i, 7, e["example_zh"])
        ws.cell(i, 8, f"images/{e['image_file']}")
        ws.cell(i, 9, e.get("exam_tags_str", ""))
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col < 6 else 40
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    print("=== 1. 解压压缩包 ===")
    staging = extract_zips()
    img_index = build_image_index(staging)

    # 归档抽出的 Excel
    CONTENT.mkdir(parents=True, exist_ok=True)
    for label, d in staging.items():
        for xp in d.glob("*.xlsx"):
            dest = CONTENT / xp.name
            safe_copy(xp, dest)
            print(f"[archive] {dest.name}")

    print("\n=== 2. 合并词表 ===")
    store = load_a1_csv()
    print(f"[info] A1 义项包 {len(store)}")

    # 基线专四（7.15 / with_images）
    base_xlsx = next((p for p in TEM4_BASE_XLSX if p.exists()), None)
    if TEM4_WITH_IMG.exists():
        for e in load_xlsx_rows(TEM4_WITH_IMG):
            merge_entry(store, e, "专四")
        print(f"[info] + tem4_with_images → {len(store)}")
    elif base_xlsx:
        for e in load_xlsx_rows(base_xlsx):
            merge_entry(store, e, "专四")
        print(f"[info] + {base_xlsx.name} → {len(store)}")

    # zip 内专四表（按日期递增覆盖）
    for label in ("tem4_716", "tem4_717", "tem4_718", "tem4_719", "tem4_720"):
        d = staging.get(label)
        if not d:
            continue
        xlsxs = list(d.glob("*.xlsx"))
        if not xlsxs:
            continue
        for e in load_xlsx_rows(xlsxs[0]):
            merge_entry(store, e, "专四")
        print(f"[info] + {xlsxs[0].name} → {len(store)}")

    # 专八
    if TEM8_WITH_IMG.exists():
        safe_copy(TEM8_WITH_IMG, CONTENT / "senses_table_tem8_with_images.xlsx")
        for e in load_xlsx_rows(TEM8_WITH_IMG):
            merge_entry(store, e, "专八")
        print(f"[info] + tem8_with_images → {len(store)}")

    embedded: dict[tuple[str, int], bytes] = {}
    if TEM4_WITH_IMG.exists():
        embedded.update(load_embedded_images(TEM4_WITH_IMG))
    if TEM8_WITH_IMG.exists():
        embedded.update(load_embedded_images(TEM8_WITH_IMG))

    print("\n=== 3. 配图落盘 ===")
    a1_dir = IMAGES / "A1"
    tem4_dir = IMAGES / "tem4"
    tem8_dir = IMAGES / "tem8"
    for d in (a1_dir, tem4_dir, tem8_dir):
        d.mkdir(parents=True, exist_ok=True)

    a1_rows: list[dict] = []
    tem4_rows: list[dict] = []
    tem8_rows: list[dict] = []
    missing_img = []

    for key, e in sorted(store.items(), key=lambda x: (x[1]["level"], x[0][0], x[0][1])):
        tags = set(e.get("exam_tags") or [])
        is_a1_pack = e.get("pack") == "A1" or "义项包" in (e.get("tags_extra") or [])
        # 分类输出：纯专八 → tem8；带专四（可同时在 A1）→ 专四 CSV 也要有一份？
        # 数据库 UNIQUE(lemma,pos,sense)，只能一行。故：
        # - 最终只写「合并后」三份 CSV，seed 时后写覆盖前写：先 A1，再 tem4，再 tem8
        # - A1 只含原先义项包键；专四含所有带专四标签；专八含专八标签
        # - 若一词同时 A1+专四，两份 CSV 都写相同内容+合并 tags，seed ON CONFLICT 合并 tags

        image_file = f"{e['image_stem']}.jpg"
        # 选择目录：优先按考试标签；A1 仅义项包装进 A1 目录
        if "专八" in tags and "专四" not in tags and not is_a1_pack:
            dest_dir = tem8_dir
        elif "专四" in tags and not is_a1_pack:
            dest_dir = tem4_dir
        elif is_a1_pack:
            dest_dir = a1_dir
        elif "专八" in tags:
            dest_dir = tem8_dir
        else:
            dest_dir = tem4_dir

        dest = dest_dir / image_file
        got = False
        if key in embedded:
            save_jpg(dest, embedded[key])
            got = True
        elif e["image_stem"] in img_index:
            save_jpg(dest, img_index[e["image_stem"]].read_bytes())
            got = True
        else:
            # 宽松：lemma_sense
            alt = f"{slugify(e['lemma'])}_{e['sense']}"
            if alt in img_index:
                image_file = f"{alt}.jpg"
                dest = dest_dir / image_file
                save_jpg(dest, img_index[alt].read_bytes())
                e["image_stem"] = alt
                got = True
        if not got:
            missing_img.append(f"{e['lemma']}#{e['sense']} ({e['image_stem']})")

        exam_tags = "|".join(sorted(tags))
        tag_parts = []
        if is_a1_pack:
            tag_parts.append("义项包")
        if tags:
            tag_parts.append("考试路径")
        tag_parts += [e["level"], e["pos"]]
        row = {
            "lemma": e["lemma"],
            "pos": e["pos"],
            "gender": e["gender"],
            "level": e["level"],
            "sense": str(e["sense"]),
            "ipa": "",
            "meaning_zh": e["meaning_zh"],
            "meaning_en": "",
            "example_es": e["example_es"],
            "example_zh": e["example_zh"],
            "image_file": image_file,
            "tags": "|".join(tag_parts),
            "exam_tags": exam_tags,
            "copyright_note": f"大创内容组_{e.get('source', 'exam')}",
            "cat_raw": e.get("cat_raw", ""),
            "exam_tags_str": exam_tags,
            "_is_a1": is_a1_pack,
            "_tags": tags,
        }
        # 分流到输出列表（可重复进入多个列表以便 seed 覆盖合并）
        if is_a1_pack:
            a1_rows.append(row)
        if "专四" in tags:
            tem4_rows.append(row)
        if "专八" in tags:
            tem8_rows.append(row)
        # 纯新词仅专四/专八已覆盖；若无考试标签且非 A1，归专四
        if not is_a1_pack and not tags:
            row["exam_tags"] = "专四"
            row["tags"] = f"考试路径|专四|{e['level']}|{e['pos']}"
            tem4_rows.append(row)

    fill_distractors(a1_rows)
    fill_distractors(tem4_rows)
    fill_distractors(tem8_rows)

    # 写 CSV
    write_csv(BATCHES / "A1" / "words_senses_team.csv", a1_rows)
    write_csv(BATCHES / "exam" / "words_tem4.csv", tem4_rows)
    write_csv(BATCHES / "exam" / "words_tem8.csv", tem8_rows)

    # 合并总表便于同学查看
    all_for_xlsx = []
    seen = set()
    for lst in (a1_rows, tem4_rows, tem8_rows):
        for r in lst:
            k = (r["lemma"], r["sense"])
            if k in seen:
                continue
            seen.add(k)
            all_for_xlsx.append(r)
    write_merged_xlsx(CONTENT / "senses_table_all_merged.xlsx", all_for_xlsx, "全部")
    write_merged_xlsx(CONTENT / "senses_table_tem4_merged.xlsx", tem4_rows, "专四")
    write_merged_xlsx(CONTENT / "senses_table_tem8_merged.xlsx", tem8_rows, "专八")

    print("\n=== 4. 汇总 ===")
    print(f"A1 CSV: {len(a1_rows)}")
    print(f"专四 CSV: {len(tem4_rows)}")
    print(f"专八 CSV: {len(tem8_rows)}")
    print(f"去重总义项: {len(all_for_xlsx)}")
    print(f"缺图: {len(missing_img)}")
    if missing_img:
        miss_path = CONTENT / "exam_missing_images.txt"
        miss_path.write_text("\n".join(missing_img) + "\n", encoding="utf-8")
        print(f"  → {miss_path}")
        for line in missing_img[:15]:
            print("  ", line)

    print("\n下一步: node scripts/reset-with-exams.mjs")


if __name__ == "__main__":
    main()
