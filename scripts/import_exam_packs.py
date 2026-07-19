#!/usr/bin/env python3
"""
导入专四 / 专八义项表 + 配图，并合并进 App 词库。

优先使用「带内嵌图」的 Excel；词条以最新 tem4 表为准。
额外 zip 配图按文件名匹配补齐；无法匹配的扩展图另存备查。

用法:
  python3 scripts/import_exam_packs.py
"""

from __future__ import annotations

import csv
import random
import re
import shutil
import sys
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CONTENT = ROOT / "data" / "content"
IMAGES_ROOT = ROOT / "data" / "images"
BATCHES = ROOT / "data" / "batches"

WECHAT_FILE = Path(
    "/Users/liuyiming/Library/Containers/com.tencent.xinWeChat/"
    "Data/Documents/xwechat_files/wxid_56yn1iomktsr12_9b4c/msg/file/2026-07"
)
WECHAT_DRAG = Path(
    "/Users/liuyiming/Library/Containers/com.tencent.xinWeChat/"
    "Data/Documents/xwechat_files/wxid_56yn1iomktsr12_9b4c/temp/drag"
)

# 最新专四词表（用户指定）
TEM4_XLSX_CANDIDATES = [
    WECHAT_DRAG / "senses_table_tem4 7.15(1).xlsx",
    WECHAT_FILE / "senses_table_tem4 7.15(1).xlsx",
    CONTENT / "senses_table_tem4.xlsx",
]
# 带内嵌配图（用于抽图）
TEM4_IMG_XLSX = WECHAT_FILE / "senses_table_tem4_with_images.xlsx"
TEM8_IMG_XLSX = WECHAT_FILE / "senses_table_tem8_with_images.xlsx"

IMAGE_ZIPS = [
    WECHAT_FILE / "images专业八级1(1).zip",
    WECHAT_FILE / "images专业四级2.zip",
    WECHAT_FILE / "images专业四级3.zip",
    WECHAT_FILE / "images专业四级4.zip",
]

CSV_FIELDS = [
    "lemma", "pos", "gender", "level", "sense", "ipa",
    "meaning_zh", "meaning_en", "example_es", "example_zh",
    "image_file", "tags", "exam_tags",
    "distractor_1", "distractor_2", "distractor_3", "copyright_note",
]

CAT_MAP = {
    "f": ("n", "f"),
    "m": ("n", "m"),
    "v": ("v", "n/a"),
    "adj": ("adj", "n/a"),
    "adv": ("adv", "n/a"),
    "interj": ("int", "n/a"),
    "int": ("int", "n/a"),
    "pron": ("pron", "n/a"),
    "prep": ("prep", "n/a"),
    "conj": ("conj", "n/a"),
}


def slugify(text: str) -> str:
    s = unicodedata.normalize("NFD", str(text).strip().lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_]+", "_", s)
    return s[:80] or "word"


def map_categoria(cat: str) -> tuple[str, str]:
    return CAT_MAP.get((cat or "").strip().lower(), ("n", "n/a"))


def extract_image_map(xlsx: Path) -> dict[int, str]:
    """Excel 1-based row → xl/media/... path inside xlsx zip."""
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    a_ns = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    r_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"

    with zipfile.ZipFile(xlsx) as z:
        rel_name = "xl/drawings/_rels/drawing1.xml.rels"
        drawing_name = "xl/drawings/drawing1.xml"
        if rel_name not in z.namelist() or drawing_name not in z.namelist():
            return {}
        rel_root = ET.fromstring(z.read(rel_name))
        rid_map = {}
        for rel in rel_root:
            if not rel.attrib.get("Type", "").endswith("/image"):
                continue
            target = rel.attrib["Target"].replace("\\", "/")
            if target.startswith("../media/"):
                target = "xl/media/" + target.split("media/", 1)[1]
            elif target.startswith("/xl/media/"):
                target = target.lstrip("/")
            elif target.startswith("media/"):
                target = "xl/" + target
            rid_map[rel.attrib["Id"]] = target
        root = ET.fromstring(z.read(drawing_name))
        mapping: dict[int, str] = {}
        for anc in root:
            frm = anc.find("xdr:from", ns)
            if frm is None:
                continue
            excel_row = int(frm.findtext("xdr:row", namespaces=ns)) + 1
            blip = anc.find(f".//{a_ns}blip")
            if blip is None:
                continue
            embed = blip.attrib.get(r_attr)
            target = rid_map.get(embed)
            if target:
                mapping[excel_row] = target
    return mapping


def load_sense_rows(xlsx: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    entries = []
    for raw in rows_raw[1:]:
        if not raw or not raw[0]:
            continue
        lemma = str(raw[0]).strip()
        sense = int(raw[1] or 1)
        cat = str(raw[2] or "").strip()
        level = str(raw[3] or "B1").strip().upper()
        meaning_zh = str(raw[4] or "").strip()
        if not meaning_zh:
            continue
        example_es = str(raw[5] or "").strip()
        example_zh = str(raw[6] or "").strip()
        image_hint = str(raw[7] or "").strip() if len(raw) > 7 else ""
        pos, gender = map_categoria(cat)
        if image_hint:
            base = Path(image_hint).name
        else:
            base = f"{slugify(lemma)}_{sense}.png"
        stem = slugify(Path(base).stem)
        ext = Path(base).suffix.lower() or ".png"
        image_file = f"{stem}{ext}"
        entries.append({
            "lemma": lemma,
            "pos": pos,
            "gender": gender,
            "level": level,
            "sense": sense,
            "meaning_zh": meaning_zh,
            "example_es": example_es,
            "example_zh": example_zh,
            "image_file": image_file,
            "image_stem": stem,
        })
    return entries


def build_zip_image_index(zips: list[Path]) -> dict[str, tuple[Path, str]]:
    """slug stem → (zip_path, member_name). Later zips override earlier."""
    index: dict[str, tuple[Path, str]] = {}
    for zp in zips:
        if not zp.exists():
            print(f"[warn] zip 不存在: {zp}")
            continue
        with zipfile.ZipFile(zp) as zf:
            for name in zf.namelist():
                if name.endswith("/") or name.startswith("__MACOSX"):
                    continue
                low = name.lower()
                if not low.endswith((".png", ".jpg", ".jpeg", ".webp")):
                    continue
                stem = slugify(Path(name).stem)
                index[stem] = (zp, name)
        print(f"[info] 索引 {zp.name}: 累计 {len(index)} 个文件名")
    return index


def extract_embedded_by_lemma(xlsx: Path) -> dict[tuple[str, int], bytes]:
    """(lemma, sense) → image bytes from embedded drawings."""
    import openpyxl

    if not xlsx.exists():
        return {}
    image_map = extract_image_map(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    out: dict[tuple[str, int], bytes] = {}
    with zipfile.ZipFile(xlsx) as z:
        for excel_row_idx, raw in enumerate(rows_raw[1:], start=2):
            if not raw or not raw[0]:
                continue
            lemma = str(raw[0]).strip()
            sense = int(raw[1] or 1)
            media = image_map.get(excel_row_idx)
            if media and media in z.namelist():
                out[(lemma, sense)] = z.read(media)
    print(f"[info] 从 {xlsx.name} 抽出内嵌图 {len(out)} 张")
    return out


def fill_distractors(entries: list[dict]) -> None:
    meanings = [e["meaning_zh"] for e in entries]
    rng = random.Random(42)
    for e in entries:
        pool = [m for m in meanings if m != e["meaning_zh"]]
        rng.shuffle(pool)
        picks = (pool + ["待填写", "待填写", "待填写"])[:3]
        e["distractor_1"], e["distractor_2"], e["distractor_3"] = picks


def write_csv(path: Path, rows: list[dict], exam_tag: str, pack_label: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out_rows = []
    for e in rows:
        out_rows.append({
            "lemma": e["lemma"],
            "pos": e["pos"],
            "gender": e["gender"],
            "level": e["level"],
            "sense": str(e["sense"]),
            "ipa": "",
            "meaning_zh": e["meaning_zh"],
            "meaning_en": "",
            "example_es": e["example_es"],
            "example_zh": e["example_zh"],
            "image_file": e["image_file"],
            "tags": f"考试路径|{exam_tag}|{e['level']}|{e['pos']}",
            "exam_tags": exam_tag,
            "distractor_1": e["distractor_1"],
            "distractor_2": e["distractor_2"],
            "distractor_3": e["distractor_3"],
            "copyright_note": f"大创内容组_{pack_label}",
        })
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(out_rows)
    print(f"[ok] CSV {path} ({len(out_rows)} 义项)")


def save_image(dest: Path, data: bytes) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)


def resolve_first_existing(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists():
            return p
    return None


def import_pack(
    *,
    name: str,
    exam_tag: str,
    vocab_xlsx: Path,
    embedded_xlsx: Path | None,
    out_csv: Path,
    out_img_dir: Path,
    zip_index: dict[str, tuple[Path, str]],
) -> list[dict]:
    entries = load_sense_rows(vocab_xlsx)
    print(f"[info] {name} 词条 {len(entries)} ← {vocab_xlsx.name}")

    embedded = extract_embedded_by_lemma(embedded_xlsx) if embedded_xlsx else {}
    # 若词表缺「intercambio académico」等，可从带图表补进
    if embedded_xlsx and embedded_xlsx.exists() and embedded_xlsx != vocab_xlsx:
        for e in load_sense_rows(embedded_xlsx):
            key = (e["lemma"], e["sense"])
            if key not in {(x["lemma"], x["sense"]) for x in entries}:
                entries.append(e)
                print(f"[info] 从带图表补入: {e['lemma']}#{e['sense']}")

    fill_distractors(entries)
    out_img_dir.mkdir(parents=True, exist_ok=True)

    got_embed = got_zip = missing = 0
    for e in entries:
        dest = out_img_dir / e["image_file"]
        key = (e["lemma"], e["sense"])
        if key in embedded:
            # 统一存为 png（内嵌多为 png）
            stem = Path(e["image_file"]).stem
            dest = out_img_dir / f"{stem}.png"
            e["image_file"] = dest.name
            save_image(dest, embedded[key])
            got_embed += 1
            continue
        hit = zip_index.get(e["image_stem"])
        if hit:
            zp, member = hit
            with zipfile.ZipFile(zp) as zf:
                data = zf.read(member)
            ext = Path(member).suffix.lower() or ".png"
            dest = out_img_dir / f"{e['image_stem']}{ext}"
            e["image_file"] = dest.name
            save_image(dest, data)
            got_zip += 1
        else:
            missing += 1
            print(f"[warn] 缺图: {e['lemma']}#{e['sense']} ({e['image_file']})")

    write_csv(out_csv, entries, exam_tag, name)
    print(f"[summary] {name}: 内嵌 {got_embed} | zip {got_zip} | 缺图 {missing}")
    return entries


def copy_extra_zip_images(
    zip_index: dict[str, tuple[Path, str]],
    used_stems: set[str],
    dest_dir: Path,
) -> int:
    """把未写入词库的扩展图存到 extras，方便同学后续补表。"""
    dest_dir.mkdir(parents=True, exist_ok=True)
    n = 0
    for stem, (zp, member) in zip_index.items():
        if stem in used_stems:
            continue
        ext = Path(member).suffix.lower() or ".png"
        out = dest_dir / f"{stem}{ext}"
        if out.exists():
            continue
        with zipfile.ZipFile(zp) as zf:
            out.write_bytes(zf.read(member))
        n += 1
    print(f"[ok] 扩展配图另存 {n} 张 → {dest_dir}")
    return n


def main() -> None:
    CONTENT.mkdir(parents=True, exist_ok=True)

    tem4_xlsx = resolve_first_existing(TEM4_XLSX_CANDIDATES)
    if not tem4_xlsx:
        raise SystemExit("找不到专四 Excel")
    # 归档到 data/content
    shutil.copy2(tem4_xlsx, CONTENT / "senses_table_tem4.xlsx")
    if TEM4_IMG_XLSX.exists():
        shutil.copy2(TEM4_IMG_XLSX, CONTENT / "senses_table_tem4_with_images.xlsx")
    if TEM8_IMG_XLSX.exists():
        shutil.copy2(TEM8_IMG_XLSX, CONTENT / "senses_table_tem8_with_images.xlsx")
        shutil.copy2(TEM8_IMG_XLSX, CONTENT / "senses_table_tem8.xlsx")

    zip_index = build_zip_image_index([p for p in IMAGE_ZIPS if p.exists()])

    tem4 = import_pack(
        name="tem4",
        exam_tag="专四",
        vocab_xlsx=CONTENT / "senses_table_tem4.xlsx",
        embedded_xlsx=CONTENT / "senses_table_tem4_with_images.xlsx"
        if (CONTENT / "senses_table_tem4_with_images.xlsx").exists()
        else TEM4_IMG_XLSX,
        out_csv=BATCHES / "exam" / "words_tem4.csv",
        out_img_dir=IMAGES_ROOT / "tem4",
        zip_index=zip_index,
    )
    tem8 = import_pack(
        name="tem8",
        exam_tag="专八",
        vocab_xlsx=CONTENT / "senses_table_tem8_with_images.xlsx"
        if (CONTENT / "senses_table_tem8_with_images.xlsx").exists()
        else CONTENT / "senses_table_tem8.xlsx",
        embedded_xlsx=CONTENT / "senses_table_tem8_with_images.xlsx"
        if (CONTENT / "senses_table_tem8_with_images.xlsx").exists()
        else TEM8_IMG_XLSX,
        out_csv=BATCHES / "exam" / "words_tem8.csv",
        out_img_dir=IMAGES_ROOT / "tem8",
        zip_index=zip_index,
    )

    used = {e["image_stem"] for e in tem4 + tem8}
    # 也把已有 A1 图文件名算作 used，避免重复巨大副本
    a1_dir = IMAGES_ROOT / "A1"
    if a1_dir.exists():
        used |= {slugify(p.stem) for p in a1_dir.iterdir() if p.is_file()}

    # 扩展图体积大，默认只索引报告，不整包拷入仓库（避免上百 MB）
    unused = [s for s in zip_index if s not in used]
    print(f"[info] zip 中未入库的扩展图 {len(unused)} 张（未整包复制，避免撑爆仓库）")
    report = CONTENT / "exam_extra_images.txt"
    report.write_text(
        "\n".join(sorted(unused)) + "\n",
        encoding="utf-8",
    )
    print(f"[ok] 扩展图清单: {report}")

    print("\n下一步:")
    print("  node scripts/reset-with-exams.mjs")


if __name__ == "__main__":
    main()
